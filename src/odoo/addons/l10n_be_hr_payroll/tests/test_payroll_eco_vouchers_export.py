# Part of Odoo. See LICENSE file for full copyright and licensing details.

import io
from datetime import date
from odoo.tests import tagged, HttpCase
from .common import TestPayrollCommon

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


@tagged('post_install_l10n', 'post_install', '-at_install')
class TestPayrollEcoVouchersExport(TestPayrollCommon, HttpCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.admin_user = cls.env.ref('base.user_admin')
        cls.admin_user.company_ids |= cls.belgian_company
        cls.admin_user.company_id = cls.belgian_company

    def test_eco_vouchers_export(self):
        self.authenticate('admin', 'admin')

        today = date.today()
        wizard = self.env['l10n.be.eco.vouchers.wizard'].with_company(self.belgian_company).create({
            'company_id': self.belgian_company.id,
            'reference_year': today.year,
        })

        self.assertTrue(wizard.line_ids, "Wizard should have lines for the employees created in common")

        response = self.url_open(f'/export/ecovouchers/{wizard.id}')
        self.assertEqual(response.status_code, 200)

        if load_workbook:
            output = io.BytesIO(response.content)
            workbook = load_workbook(output)
            sheet = workbook['Worksheet']

            expected_headers = [
                "Numéro de registre national (p.ex. 790227 183 12)",
                "Salarié nom (p.ex. dupont)",
                "Salarié prénom (p.ex. max)",
                "Votre numéro interne du salarié  (p.ex. 152d97)",
                "Nombre de chèques [a] (p.ex. 18)",
                "Valeur faciale du chèque [b] (p.ex. 5.5)",
                "Total [a] x [b] (p.ex. 99)",
                "Date de naissance du salarié (dd/mm/yyyy)",
                "Sexe du salarié (m/f)",
                "Langue du salarié (nl/fr/en)",
                "Centre de coûts (p.ex. cc1. maximum 10 caractères)",
                "Votre numéro d'entreprise  (p.ex. be 0834013324)",
                "Adresse de livraison rue (p.ex. av. des volontaires)",
                "Adresse de livraison numéro (p.ex. 19)",
                "Adresse de livraison boite (p.ex. a1)",
                "Adresse de livraison code postal (p.ex. 1160)",
                "Adresse de livraison ville (p.ex. auderghem)",
                "Statut contrat",
            ]

            actual_headers = [cell.value for cell in sheet[1]]
            self.assertEqual(actual_headers, expected_headers)

            self.assertTrue(sheet.max_row > 1, "There should be at least one data row")

            # Check the 'Statut contrat' column
            last_col_index = len(expected_headers)
            status_value = sheet.cell(row=2, column=last_col_index).value
            self.assertIn(status_value, ['Actif', 'Fin de la collaboration'])
